from tkinter import *

from tkinter import messagebox
from docxtpl import DocxTemplate
from tkdocviewer import *
# from win32com.client import constants, Dispatch
# from win32com.gen_py.word import *
# py -3.12 -m pip install docxtpl
import datetime

import customtkinter
# py -3.12 -m pip install customtkinter
import openpyxl
# pip install openpyxl
import tkinter
# main frame
import tkinter as tk
from tkinter import ttk
root = Tk()
root.geometry("1000x700")
root.resizable(width = False, height= False)
# root.configure(bg="#153448")

style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

root.title('Zaman Medical Store')
usr = "blank user"
label = ttk.Label(text="Zaman Medical Store", font='System 25 bold', foreground="#9BEC00")
label.place(relx=0.5, rely=0.3, anchor=CENTER)

fexo = 100
ometid = 200
montair = 500
saline = 300
napa = 500
tofen = 100
azicin = 200
vargon = 100
# Login Function

def login_func():
    label.destroy()
    login.destroy()
    register.destroy()

    label2 = Label(text="Log In", font='System 33 bold', fg="#9BEC00")
    label2.place(relx=0.5, rely=0.2, anchor=CENTER)

    name = customtkinter.CTkLabel(root, text="Username :", font=("System", 30), corner_radius=50)
    name.place(relx=0.4, rely=0.4, anchor=CENTER)

    name_entry = customtkinter.CTkEntry(root, placeholder_text="Enter your username")
    name_entry.place(relx=0.6, rely=0.4, anchor=CENTER)
    
    
    password = customtkinter.CTkLabel(root, text="Password  :", font=("System", 30), corner_radius=50)
    password.place(relx=0.4, rely=0.5, anchor=CENTER)

    password_entry = customtkinter.CTkEntry(root, placeholder_text="Enter your Password", show='*')
    password_entry.place(relx=0.6, rely=0.5, anchor=CENTER)

    button = ttk.Button(root, text="Enter", width=25, command=lambda: check(name_entry.get() + " " + password_entry.get()))
    button.place(relx=0.5, rely=0.65, anchor=CENTER)
    user_name = name_entry.get()


    # to check the password and the username

    def check(info):
        f = open("admin_info.txt", "a")
        g = open("admin_info.txt", "r")
        user = info.split()
        flag = 0
        for x in g:
            y = x.split()
            if y[0] == user[0] and y[1] == user[1]:
                # messagebox.showwarning("Welcome", "Login Successful")
                label2.destroy()
                button.destroy()
                name.destroy()
                name_entry.destroy()
                password.destroy()
                password_entry.destroy()
                flag = 1
                home()
                break
        if flag == 0:
            messagebox.showwarning("Sorry", "Wrong Credentials.")
            label2.destroy()
            name.destroy()
            name_entry.destroy()
            password.destroy()
            password_entry.destroy()
            button.destroy()
            login_func()
        f.close()
        g.close()
    
    # home page function
    
    def home():
        label2.destroy()
        button.destroy()
        name.destroy()
        name_entry.destroy()
        password.destroy()
        password_entry.destroy()
        
        home_frame = tkinter.Frame(root)
        home_frame.pack()
        buttons_frame = ttk.LabelFrame(home_frame)
        buttons_frame.grid(row=0,column=0, padx= 20, pady = 10)

        # buy and invoice
        def invoice():
            home_frame.destroy()
            buttons_frame.destroy()
           # invoice_label = Label(text="Invoice", font='System 28 bold', bg="#153448", fg="#DFD0B8")
           # invoice_label.place(relx=0.5, rely=0.1, anchor=CENTER)
            values_product_name = []
            path = "Product_list.xlsx"
            wb = openpyxl.load_workbook(path)
            sheet = wb['Sheet1']
            row_count = sheet.max_row
            for cellObj in sheet.iter_rows(min_row=2, min_col=1,max_row=row_count, max_col=1):
                for cell in cellObj:
                    values_product_name.append(str(cell.value))
                    
            frame = tkinter.Frame(root)
            frame.pack()

            first_name_label = ttk.Label(frame, text = "First Name")
            first_name_label.grid(row =2, column=0)
            last_name_label = ttk.Label(frame, text="Last Name")
            last_name_label.grid(row=2, column=1)

            first_name_entry = ttk.Entry(frame)
            first_name_entry.grid(row =3, column=0)
            first_name_entry.insert(0, user_name)
            last_name_entry = customtkinter.CTkEntry(frame)
            last_name_entry.grid(row=3, column=1)

            phone_label = ttk.Label(frame, text="Phone")
            phone_label.grid(row = 2, column = 2)
            phone_entry = customtkinter.CTkEntry(frame)
            phone_entry.grid(row=3,column=2)

            qty_label = ttk.Label(frame, text="Quantity")
            qty_label.grid(row=4,column=0)
            qty_spinbox = customtkinter.CTkEntry(frame)
            qty_spinbox.grid(row = 5, column= 0)

            desc_label = ttk.Label(frame, text="Description")
            desc_label.grid(row=4, column=1)
            desc_entry = ttk.Combobox(frame, values= values_product_name)
            desc_entry.current(0)
            desc_entry.grid(row = 5, column= 1)

            price_label = ttk.Label(frame, text="Unit Price")
            price_label.grid(row = 4, column= 2)
            price_spinbox = customtkinter.CTkEntry(frame)
            # price_spinbox = tkinter.Spinbox(frame, from_=0.00, to=500.00, increment= 0.5)
            price_spinbox.grid(row = 5, column = 2)
             
            # generate invoice
            def generate_invoice():
                doc = DocxTemplate("invoice_template.docx")
                name = first_name_entry.get()+" "+last_name_entry.get()
                phone = phone_entry.get()
                subtotal = sum(item[3] for item in invoice_list)
                salestax = 0.1
                total = subtotal*(1-salestax)

                doc.render({"name":name,
                            "phone":phone,
                            "invoice_list":invoice_list,
                            "subtotal":subtotal,
                            "salestax":str(salestax*100)+ "%",
                            "total":total})
                doc_name = "new_invoice"+name+datetime.datetime.now().strftime("%Y-%m-%d-%H%M%S")+".docs"
                doc.save(doc_name)
                
                doc_window = Toplevel(root)
                doc_window.title("Invoice")
                doc_window.geometry("1000x1000")
                doc_window.resizable(width= True, height=False)
                v1 = DocViewer(doc_window)
                v1.pack(side="top", expand=1, fill="both")
                v1.display_file(doc_name)
                messagebox.showinfo("invoice complete", "Invoice Created")

                new_invoice()


            # new invoice function notun invoice create korbe
            def new_invoice():
                first_name_entry.delete(0,tkinter.END)
                last_name_entry.delete(0, tkinter.END)
                phone_entry.delete(0,tkinter.END)
                clear_items()
                tree.delete(*tree.get_children())

                invoice_list.clear()

            # function to clear the items
            def clear_items():
                qty_spinbox.delete(0,tkinter.END)
                qty_spinbox.insert(0, "")
                desc_entry.delete(0, tkinter.END)
                price_spinbox.delete(0, tkinter.END)
                price_spinbox.insert(0,"")

            invoice_list = []
            # add func to add all the info to the invoice
            def add_item():
                qty = int(qty_spinbox.get())
                desc = desc_entry.get()
                price = float(price_spinbox.get())
                line_total = qty*price

                invoice_item = [qty, desc,price,line_total]

                tree.insert('', 0, values=invoice_item)
                clear_items()

                invoice_list.append(invoice_item)

            add_item_button = ttk.Button(frame, text= "Add to the Cart", command=add_item)
            add_item_button.grid(row=6, column = 2, pady=5)

            columns = ('qty', 'desc', 'price','total')
            tree = ttk.Treeview(frame, columns=columns, show="headings")
            tree.heading('qty',text="Qty")
            tree.heading('desc', text = "Description")
            tree.heading('price', text= "Price")
            tree.heading('total', text="Total")
            tree.grid(row=7, column= 0, columnspan=3, padx=20,pady=10)

            save_invoice_button = ttk.Button(frame, text= "Generate Invoice", command=generate_invoice)
            save_invoice_button.grid(row=8, column = 0, columnspan=3, sticky="news", padx=20, pady=5)

            new_invoice_button = ttk.Button(frame, text= "New Invoice", command=new_invoice)
            new_invoice_button.grid(row=9, column=0, columnspan = 3, sticky= "news", padx= 20, pady=5)

        buy = ttk.Button(buttons_frame,text="Buy",command=invoice)
        buy.grid(row=0,column=0, padx=5, pady=5, sticky="ew")

        # add products function

        def add_p():
            home_frame.destroy()
            buttons_frame.destroy()

            frame = tkinter.Frame(root)
            frame.pack()

            widgets_frame = ttk.LabelFrame(frame, text="Insert Product")
            widgets_frame.grid(row=0,column=0, padx= 20, pady = 10)

            new_product_name_entry = ttk.Entry(widgets_frame)
            new_product_name_entry.insert(0,"Enter The Product Name")
            new_product_name_entry.bind("<FocusIn>", lambda e: new_product_name_entry.delete('0','end'))
            new_product_name_entry.grid(row=0,column=0,padx=5, pady= (0,5),sticky="ew")

            quantity_spinbox = ttk.Spinbox(widgets_frame, from_ = 1, to = 100)
            quantity_spinbox.grid(row = 1, column=0, padx=5, pady=5, sticky="ew")
            quantity_spinbox.insert(0,"Enter Quantity")
            quantity_spinbox.bind("<FocusIn>", lambda e: quantity_spinbox.delete('0','end'))

            added_time = datetime.datetime.now().strftime("%Y-%m-%d")
            added_time_entry = ttk.Entry(widgets_frame)
            added_time_entry.insert(0, added_time)
            added_time_entry.grid(row=2,column=0, padx=5, pady=5, sticky="ew")

            # function behind insert button
            def insert_product_func():
                newproductname = new_product_name_entry.get()
                newproductquantity = int(quantity_spinbox.get())
                product_entry_date = added_time_entry.get()

                path = "Product_list.xlsx"
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active
                row_values = [newproductname, newproductquantity, product_entry_date]
                sheet.append(row_values)
                workbook.save(path)
                treeview.insert('',tk.END, values=row_values)
                # clear the values
                new_product_name_entry.delete(0,"end")
                new_product_name_entry.insert(0,"Enter The Product Name")
                quantity_spinbox.delete(0, "end")
                quantity_spinbox.insert(0, "Enter Quantity")



            insert_button = ttk.Button(widgets_frame, text="Insert", command= insert_product_func)
            insert_button.grid(row=3,column=0, sticky="ew")

            separator = ttk.Separator(widgets_frame)
            separator.grid(row=4,column=0,padx=(20,10),pady=10,sticky="ew")
            # toggle function (day/night mode changer)
            def toggle_mode():
                if mode_switch.instate(["selected"]):
                    style.theme_use("forest-light")
                else:
                    style.theme_use("forest-dark")
            # day-night mode switcher
            mode_switch = ttk.Checkbutton(widgets_frame,text="Day/Night Mode", style="Switch", command=toggle_mode)
            mode_switch.grid(row = 5, column = 0, padx=5, pady = 10, sticky="ew")
            
            # function to show data from excel file is here
            def load_data():
                path = "C:/Users/Acer/OneDrive/Documents/GitHub/Medical-Store-Management/Product_list.xlsx"
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active
                list_values = list(sheet.values)
                print(list_values)
                for col_name in list_values[0]:
                    treeview.heading(col_name, text=col_name)
                for value_tuple in list_values[1:]:
                    treeview.insert('', tk.END, values=value_tuple)   

            treeframe = ttk.Frame(frame)
            treeframe.grid(row=0, column=1, pady=10)
            treescroll = ttk.Scrollbar(treeframe)
            treescroll.pack(side="right", fill="y")

            cols=("Product Name", "Quantity", "Date")
            treeview = ttk.Treeview(treeframe, show="headings",yscrollcommand= treescroll.set,columns = cols, height=13)
            # columns
            treeview.column("Product Name", width=200)
            treeview.column("Quantity", width=100)
            treeview.column("Date", width=200)
            treeview.pack()
            treescroll.config(command = treeview.yview)
            load_data()

        add = ttk.Button(buttons_frame,text="Add",command=add_p)
        add.grid(row=1,column=0,padx=5, pady=5 ,sticky="ew")
        search = ttk.Button(buttons_frame,text="Search")
        search.grid(row=2,column=0,padx=5, pady=5, sticky="ew")

        def l():  # list function
            """
            p = Toplevel(root)
            p.title("Product List")
            p.geometry("1000x600")
            label_p = Label(p, text="Product List", font='System 28 bold', bg="#153448", fg="#DFD0B8")
            label_p.place(relx=0.5, rely=0.1, anchor=CENTER)
            p.resizable(width=False, height=False)

            """
            buy.destroy()
            add.destroy()
            search.destroy()
            lll.destroy()
            # p.configure(bg="#153448")
            # this is a tree
            product_list = ttk.Treeview(root)
            # label of product list
            label3 = Label(text="Product List", font='System 28 bold', bg="#153448", fg="#DFD0B8")
            label3.place(relx=0.5, rely=0.2, anchor=CENTER)
            # define our columns
            product_list['columns'] = ("Item No.", "Name", "Quantity")
            # format our columns
            product_list.column("#0", width=120, minwidth=25)
            product_list.column("Item No.", anchor=W, width=120, minwidth=25)
            product_list.column("Name", anchor=CENTER, width=120)
            product_list.column("Quantity", anchor=W, width=120)

            # Create Headings
            product_list.heading("#0", text="Label", anchor=W)
            product_list.heading("Item No.", text="Item No.", anchor=W)
            product_list.heading("Name", text="Name", anchor=CENTER)
            product_list.heading("Quantity", text="Quantity", anchor=W)

            product_list.insert(parent='', index='end', iid=0, values=(1, "Fexofenadine", fexo))
            product_list.insert(parent='', index='end', iid=1, values=(2, "Ometidine", ometid))
            product_list.insert(parent='', index='end', iid=2, values=(3, "Montair", montair))
            product_list.insert(parent='', index='end', iid=4, values=(4, "Orsaline", saline))
            product_list.insert(parent='', index='end', iid=5, values=(5, "Napa-Paracetamol", napa))
            product_list.insert(parent='', index='end', iid=6, values=(6, "Tofen", tofen))
            product_list.insert(parent='', index='end', iid=7, values=(7, "Azicin-Antibiotic", azicin))
            product_list.insert(parent='', index='end', iid=8, values=(8, "vargon", vargon))
            product_list.insert(parent='', index='end', iid=9, values=(9, "vargon", vargon))
            product_list.insert(parent='', index='end', iid=10, values=(10, "vargon", vargon))
            product_list.place(relx=0.5, rely=0.5, anchor=CENTER)
            style = ttk.Style(root)
            style.configure("Treeview", background='#153448', foreground='#DFD0B8')
            
        lll = ttk.Button(buttons_frame,text="Product List", command=l)
        lll.grid(row=3,column=0,padx=5, pady=5,sticky="ew")

        def load_data_home_menu():
           # path = "F:/BIE/Python/Medical Store Management/Product_list.xlsx"
            path = "C:/Users/Acer/OneDrive/Documents/GitHub/Medical-Store-Management/Product_list.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            list_values = list(sheet.values)
            # print(list_values)
            for col_name in list_values[0]:
                treeview.heading(col_name, text=col_name)
            for value_tuple in list_values[1:]:
                treeview.insert('', tk.END, values=value_tuple)   

        treeframe = ttk.Frame(home_frame)
        treeframe.grid(row=0, column=1, pady=10)
        treescroll = ttk.Scrollbar(treeframe)
        treescroll.pack(side="right", fill="y")

        cols=("Product Name", "Quantity", "Date")
        treeview = ttk.Treeview(treeframe, show="headings",yscrollcommand= treescroll.set,columns = cols, height=13)
        # columns
        treeview.column("Product Name", width=200)
        treeview.column("Quantity", width=100)
        treeview.column("Date", width=200)
        treeview.pack()
        treescroll.config(command = treeview.yview)
        load_data_home_menu()

        def toggle_mode_home_menu():
            if mode_switch.instate(["selected"]):
                style.theme_use("forest-light")
            else:
                style.theme_use("forest-dark")
            # day-night mode switcher
        mode_switch = ttk.Checkbutton(buttons_frame,text="Day/Night Mode", style="Switch", command=toggle_mode_home_menu)
        mode_switch.grid(row = 4, column = 0, padx=5, pady = 10, sticky="ew")

# Register Page

def reg_func():
    label.destroy()
    login.destroy()
    register.destroy()
    label2 = Label(text="Sign Up", font='System 33 bold', bg="#153448", fg="#DFD0B8")
    label2.place(relx=0.5, rely=0.2, anchor=CENTER)

    name = customtkinter.CTkLabel(root, text="Username :", font=("System", 30), corner_radius=50)
    name.place(relx=0.4, rely=0.4, anchor=CENTER)

    name_entry = customtkinter.CTkEntry(root, placeholder_text="Enter your username")
    name_entry.place(relx=0.6, rely=0.4, anchor=CENTER)

    password = customtkinter.CTkLabel(root, text="Password  :", font=("System", 30), corner_radius=50)
    password.place(relx=0.4, rely=0.5, anchor=CENTER)

    password_entry = customtkinter.CTkEntry(root, placeholder_text="Enter your Password")
    password_entry.place(relx=0.6, rely=0.5, anchor=CENTER)

    button = ttk.Button(root, text="Register", width=25,command=lambda : submit(name_entry.get()+" "+password_entry.get()))
    button.place(relx=0.5, rely=0.65, anchor=CENTER)

    # sign up submit button function

    def submit(info):
        f = open("admin_info.txt", "a")
        g = open("admin_info.txt", "r")
        user = info.split()
        for x in g:
            y = x.split()
            if y[0] == user[0]:
                messagebox.showwarning("showwarning", "User Already Registered!!!")
                f.close()
                reg_func()
            else:
                label2.destroy()
                name.destroy()
                name_entry.destroy()
                password.destroy()
                password_entry.destroy()
                button.destroy()
                login_func()
        f.write(info+"\n")
        f.close()
        g.close()

# Login button
login = ttk.Button (text="Log in", width=20, command=login_func)
login.place(relx=0.5, rely=0.5, anchor=CENTER)

# Register Button
register = ttk.Button (text="Sign Up", width=20,command=reg_func)
register.place(relx=0.5, rely=0.55, anchor=CENTER)

root.mainloop()